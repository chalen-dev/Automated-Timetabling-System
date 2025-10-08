# ===============================================
# AUTOMATED TIMETABLING SCRIPT (COLAB READY)
# ===============================================
# Assumes all required CSVs are already uploaded to /content/
# course-sessions.csv
# session-groups.csv
# timetable-rooms.csv
# timetable_template.csv
# (professors CSV ignored for now)

import pandas as pd
import numpy as np
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

# -----------------------------------------------
# ARGUMENTS FROM LARAVEL
# -----------------------------------------------
if len(sys.argv) < 4:
    print("Usage: python process_timetable.py <inputDir> <outputDir> <timetableId>")
    sys.exit(1)

input_dir = sys.argv[1]       # Path where CSVs are stored (from Laravel)
output_dir = sys.argv[2]      # Path where XLSX should be saved
timetable_id = sys.argv[3]    # Used for output filename

print(f"Running timetable generation for timetable ID: {timetable_id}")
print(f"Using CSV input directory: {input_dir}")
print(f"XLSX will be saved to: {output_dir}")

# -----------------------------------------------
# CONFIG / INPUT PATHS
# -----------------------------------------------
course_sessions_path    = os.path.join(input_dir, "course-sessions.csv")
session_groups_path     = os.path.join(input_dir, "session-groups.csv")
timetable_rooms_path    = os.path.join(input_dir, "timetable-rooms.csv")
timetable_template_path = os.path.join(input_dir, "timetable_template.csv")


# -----------------------------------------------
# LOAD DATA
# -----------------------------------------------
course_sessions = pd.read_csv(course_sessions_path)
session_groups = pd.read_csv(session_groups_path)
timetable_rooms = pd.read_csv(timetable_rooms_path)
timetable_template = pd.read_csv(timetable_template_path)

# -----------------------------------------------
# BASIC SANITY CHECKS
# -----------------------------------------------
print("Loaded files:")
for name, df in {
    "course_sessions": course_sessions,
    "session_groups": session_groups,
    "timetable_rooms": timetable_rooms,
    "timetable_template": timetable_template,
}.items():
    print(f"  {name:<20}: {df.shape} rows={len(df)}")

# -----------------------------------------------
# FILTER AND PREP
# -----------------------------------------------
# normalize room exclusivity
timetable_rooms["course_type_exclusive_to"] = timetable_rooms["course_type_exclusive_to"].fillna("none").str.lower()
timetable_rooms["room_type"] = timetable_rooms["room_type"].str.lower()

# Merge course_sessions with their session_group info
merged = course_sessions.merge(session_groups, on=["session_group_id", "year_level", "academic_program"], how="left")

# -----------------------------------------------
# ASSIGN ROOMS BASED ON TYPE + EXCLUSIVITY RULES
# -----------------------------------------------
def assign_room(row, rooms):
    ctype = row["course_type"].lower()
    rtype = "lecture" if row["total_laboratory_class_days"] == 0 else "comlab"
    # filter by exclusivity rules
    valid_rooms = rooms[
        ((rooms["course_type_exclusive_to"] == "none") & (~rooms["room_type"].isin(["pe", "nstp"])))
        | (rooms["course_type_exclusive_to"] == ctype)
    ]
    # prefer matching room_type
    valid_rooms = valid_rooms[valid_rooms["room_type"].str.contains(rtype, case=False, na=False)]
    if len(valid_rooms) == 0:
        return np.nan
    return valid_rooms.sample(1).iloc[0]["room_name"]

merged["assigned_room"] = merged.apply(assign_room, rooms=timetable_rooms, axis=1)

# -----------------------------------------------
# TIME GENERATION RULES (7:00–18:00, 30-min steps)
# -----------------------------------------------
time_slots = [f"{h:02d}:{m:02d}" for h in range(7, 18) for m in (0, 30)]
time_slots = [t for t in time_slots if not ("12:00" <= t < "12:30")]  # enforce lunch break

# -----------------------------------------------
# SIMPLE GREEDY TIMETABLING (per term)
# -----------------------------------------------
def generate_timetable(term_df, rooms, template):
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    result = {d: template.copy(deep=True) for d in days}

    # iterate per session_group
    for sg_id, group_df in term_df.groupby("session_group_id"):
        for _, cs in group_df.iterrows():
            hours = cs["class_hours"]
            room = cs["assigned_room"]
            if pd.isna(room):
                continue
            slot_len = int(hours * 2)  # 2 slots/hour (30min each)
            placed = False
            for day in days:
                for i in range(len(time_slots) - slot_len):
                    block = time_slots[i : i + slot_len]
                    if all(result[day].loc[i + j, room] == "vacant" for j in range(slot_len)):
                        for j in range(slot_len):
                            result[day].loc[i + j, room] = f"{cs['program_abbreviation']}_{cs['year_level']}_{cs['session_group_id']}_{cs['course_session_id']}"
                        placed = True
                        break
                if placed:
                    break
    return result

# Separate by term
term1_df = merged[merged["academic_term"].str.contains("1st", case=False, na=False)]
term2_df = merged[merged["academic_term"].str.contains("2nd", case=False, na=False)]
sem_df  = merged[merged["academic_term"].str.contains("sem", case=False, na=False)]

# Apply semestral rule: duplicate semestral courses to both terms
term1_df = pd.concat([term1_df, sem_df], ignore_index=True)
term2_df = pd.concat([term2_df, sem_df], ignore_index=True)

timetable_term1 = generate_timetable(term1_df, timetable_rooms, timetable_template.copy())
timetable_term2 = generate_timetable(term2_df, timetable_rooms, timetable_template.copy())

# -----------------------------------------------
# CREATE XLSX OUTPUT yeah
# -----------------------------------------------
wb = Workbook()
wb.remove(wb.active)

def add_timetable_sheets(wb, term_timetable, term_label):
    for day, df in term_timetable.items():
        ws = wb.create_sheet(f"{term_label}_{day}")
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

add_timetable_sheets(wb, timetable_term1, "1stTerm")
add_timetable_sheets(wb, timetable_term2, "2ndTerm")

# COURSE_SESSION OVERVIEW SHEETS
def make_overview(df, term_label):
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    overview = pd.DataFrame({"course_session": []})
    overview["course_session"] = df.apply(lambda x: f"{x['program_abbreviation']}_{x['year_level']}_{x['session_group_id']}_{x['course_session_id']}", axis=1)
    for day in days:
        overview[day] = df["assigned_room"].fillna("")
    return overview

overview1 = make_overview(term1_df, "1stTerm")
overview2 = make_overview(term2_df, "2ndTerm")

for name, data in [("Overview_1stTerm", overview1), ("Overview_2ndTerm", overview2)]:
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

# UNASSIGNED
unassigned = merged[merged["assigned_room"].isna()]
if not unassigned.empty:
    ws = wb.create_sheet("Unassigned")
    for r in dataframe_to_rows(unassigned, index=False, header=True):
        ws.append(r)

# Define full XLSX path
output_path = os.path.join(output_dir, f"{timetable_id}.xlsx")

wb.save(output_path)
print(f"\n Timetable successfully generated and saved as {output_path}")
